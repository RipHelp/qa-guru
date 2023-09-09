from openpyxl import load_workbook
from os import path
from conftest import RESOURCES_DIR

# оформить в тест, добавить ассерты и использовать универсальный путь
my_xlsx = path.join(RESOURCES_DIR, "file_example_XLSX_50.xlsx")
names = []
actual_names = ['Dulce', 'Mara', 'Philip', 'Kathleen', 'Nereida', 'Gaston']


def test_read_xlsx():
    workbook = load_workbook(my_xlsx)

    assert workbook.sheetnames[0] == 'Sheet1'

    sheet = workbook.active
    result = sheet.cell(row=3, column=2).value

    assert result == 'Mara'
    assert path.getsize(my_xlsx) == 7360